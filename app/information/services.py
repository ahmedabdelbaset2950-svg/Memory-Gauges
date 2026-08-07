from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.models.information import InformationRow
import pandas as pd

from app.extensions import db
from app.models.information import (
    InformationMonth,
    InformationRow
)


# ==========================================================
# Get Next Group Number
# ==========================================================

def get_next_group_no(year, month):

    last_row = (
        InformationRow.query
        .filter_by(
            year=year,
            month=month
        )
        .order_by(
            InformationRow.group_no.desc()
        )
        .first()
    )

    if not last_row:
        return 1

    return last_row.group_no + 1


# ==========================================================
# Build Information Excel Workbook
# ==========================================================

def build_information_workbook(year, month):

    rows = (
        InformationRow.query
        .filter_by(
            year=year,
            month=month
        )
        .order_by(
            InformationRow.group_no.asc(),
            InformationRow.id.asc()
        )
        .all()
    )

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Information"

    # ======================================================
    # Headers
    # ======================================================

    headers = [
        "No",
        "Gauge S/N",
        "From",
        "To",
        "Days",
        "Well Number",
        "Changed To",
        "Survey",
        "Type",
        "Rig Name",
        "Bundle Carrier S/N",
        "Battery S/N",
        "Engineer",
        "Total Hours",
        "Comment"
    ]

    for column_number, header in enumerate(headers, start=1):

        cell = sheet.cell(
            row=1,
            column=column_number,
            value=header
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="173B68"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # ======================================================
    # Data
    # ======================================================

    for row_number, item in enumerate(rows, start=2):

        values = [
            item.group_no,
            item.gauge_serial,
            item.from_date,
            item.to_date,
            item.days,
            item.well_number,
            item.changed_to,
            item.survey,
            item.type,
            item.rig_name,
            item.bundle_carrier_sn,
            item.battery_sn,
            item.engineer,
            item.total_hours,
            item.comment
        ]

        for column_number, value in enumerate(values, start=1):

            cell = sheet.cell(
                row=row_number,
                column=column_number,
                value=value
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    # ======================================================
    # Borders
    # ======================================================

    thin_border = Side(
        style="thin",
        color="D9E2F3"
    )

    for row in sheet.iter_rows():

        for cell in row:

            cell.border = Border(
                left=thin_border,
                right=thin_border,
                top=thin_border,
                bottom=thin_border
            )

    # ======================================================
    # Column Widths
    # ======================================================

    widths = {
        "A": 8,
        "B": 18,
        "C": 14,
        "D": 14,
        "E": 10,
        "F": 18,
        "G": 18,
        "H": 16,
        "I": 16,
        "J": 18,
        "K": 22,
        "L": 18,
        "M": 20,
        "N": 15,
        "O": 30
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"

    return workbook
import pandas as pd
from app.extensions import db
from app.models.information import InformationMonth, InformationRow


def import_information_excel(file):

    excel = pd.ExcelFile(file)

    imported = 0

    REQUIRED_COLUMNS = [
        "No",
        "Gauge Type S/N (Used )",
        "From",
        "To",
        "No(D)",
        "Well No",
        "Changed To",
        "Survey",
        "Type",
        "Rig Name",
        "B.Carrier S/N",
        "Battery S/N",
        "ENGINEER",
        "Total Used (H)",
        "Comment"
    ]

    for sheet_name in excel.sheet_names:

        print(f"\n========== {sheet_name} ==========")

        try:

            df = pd.read_excel(
                excel,
                sheet_name=sheet_name,
                header=8
            )

            df.columns = (
                df.columns
                .astype(str)
                .str.strip()
            )

            if df.empty:
                print("Sheet Empty")
                continue

            missing = [
                col for col in REQUIRED_COLUMNS
                if col not in df.columns
            ]

            if missing:
                print(f"Missing Columns : {missing}")
                continue

            df = df[
                df["Gauge Type S/N (Used )"]
                .notna()
            ].copy()

            if df.empty:
                continue

            first_date = pd.to_datetime(
                df.iloc[0]["From"],
                errors="coerce"
            )

            if pd.isna(first_date):
                print("Invalid Date")
                continue

            year = first_date.year
            month = first_date.month

            print(f"Importing {month}/{year}")

            InformationRow.query.filter_by(
                year=year,
                month=month
            ).delete()

            month_obj = InformationMonth.query.filter_by(
                year=year,
                month=month
            ).first()

            if not month_obj:

                month_obj = InformationMonth(
                    year=year,
                    month=month
                )

                db.session.add(month_obj)

            current_group = 0

            for _, row in df.iterrows():

                if pd.notna(row["No"]):
                    current_group = int(row["No"])

                from_date = pd.to_datetime(
                    row["From"],
                    errors="coerce"
                )

                to_date = pd.to_datetime(
                    row["To"],
                    errors="coerce"
                )

                info = InformationRow(

                    year=year,
                    month=month,

                    group_no=current_group,

                    gauge_serial=str(
                        row["Gauge Type S/N (Used )"]
                    ).strip(),

                    from_date=from_date.date()
                    if pd.notna(from_date)
                    else None,

                    to_date=to_date.date()
                    if pd.notna(to_date)
                    else None,

                    days=float(row["No(D)"])
                    if pd.notna(row["No(D)"])
                    else 0,

                    well_number=str(
                        row["Well No"]
                    ).strip(),

                    changed_to=str(
                        row["Changed To"]
                    ).strip(),

                    survey=str(
                        row["Survey"]
                    ).strip(),

                    type=str(
                        row["Type"]
                    ).strip(),

                    rig_name=str(
                        row["Rig Name"]
                    ).strip(),

                    position="",

                    bundle_carrier_sn=str(
                        row["B.Carrier S/N"]
                    ).strip(),

                    battery_sn=str(
                        row["Battery S/N"]
                    ).strip(),

                    engineer=str(
                        row["ENGINEER"]
                    ).strip(),

                    total_hours=float(
                        row["Total Used (H)"]
                    ) if pd.notna(row["Total Used (H)"]) else 0,

                    total_samples=0,

                    comment="" if pd.isna(row["Comment"])
                    else str(row["Comment"]).strip()

                )

                db.session.add(info)
                imported += 1

            db.session.commit()

            print(f"Imported {sheet_name} Successfully")

        except Exception as e:

            db.session.rollback()

            print(f"ERROR IN SHEET {sheet_name}")
            print(e)

            continue

    return imported
