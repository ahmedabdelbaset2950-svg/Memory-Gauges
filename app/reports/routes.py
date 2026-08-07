import io

from collections import defaultdict
from datetime import date, datetime

from flask import (
    Blueprint, render_template, request, redirect, url_for, flash, send_file
)
from flask_login import login_required
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.utils import get_column_letter

from openpyxl.chart import (
    BarChart,
    PieChart,
    LineChart,
    Reference
)

from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo
)

from openpyxl.formatting.rule import (
    ColorScaleRule
)

from openpyxl.worksheet.page import (
    PageMargins
)

from app.extensions import db
from app.models.information import InformationRow
from app.models.maintenance import MaintenanceRecord
from app.models.memory_gauge import MemoryGauge
from app.models.bundle_carrier import BundleCarrier
from app.models.battery import Battery
from app.utils.excel_report import ExcelReport


reports = Blueprint("reports", __name__, url_prefix="/reports")
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="173B68")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
ZEBRA_FILL = PatternFill("solid", fgColor="F4F7FB")
THIN = Side(style="thin", color="D9E1EA")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FONT = Font(bold=True, size=14, color="173B68")
SUBTITLE_FONT = Font(italic=True, size=10, color="6B7280")
TITLE_FILL = PatternFill(
    "solid",
    fgColor="173B68"
)

INFO_FILL = PatternFill(
    "solid",
    fgColor="EAF2FB"
)

WHITE_FONT = Font(
    bold=True,
    color="FFFFFF",
    size=18
)

INFO_FONT = Font(
    bold=True,
    color="173B68",
    size=11
)

TITLE_CENTER = Alignment(
    horizontal="center",
    vertical="center"
)


def _write_sheet(
    ws,
    headers,
    rows,
    title=None,
    subtitle=None,
    number_cols=None
):

    number_cols = number_cols or set()

    row_offset = 0

    # =====================================================
    # EXECUTIVE BANNER
    # =====================================================

    if title:

        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=2,
            end_column=len(headers)
        )

        banner = ws["A1"]

        banner.value = "MGMS • Memory Gauge Management System"

        banner.font = WHITE_FONT

        banner.fill = TITLE_FILL

        banner.alignment = TITLE_CENTER

        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 8

        ws.merge_cells(
            start_row=4,
            start_column=1,
            end_row=4,
            end_column=len(headers)
        )

        title_cell = ws["A4"]

        title_cell.value = title

        title_cell.font = TITLE_FONT

        title_cell.alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        if subtitle:

            ws.merge_cells(
                start_row=5,
                start_column=1,
                end_row=5,
                end_column=len(headers)
            )

            sub = ws["A5"]

            sub.value = subtitle

            sub.font = SUBTITLE_FONT

            sub.alignment = LEFT

        row_offset = 6

    # =====================================================
    # HEADER
    # =====================================================

    header_row = row_offset + 1

    for col_idx, text in enumerate(headers, start=1):

        cell = ws.cell(
            row=header_row,
            column=col_idx,
            value=text
        )

        cell.font = HEADER_FONT

        cell.fill = HEADER_FILL

        cell.alignment = CENTER

        cell.border = CELL_BORDER

    ws.row_dimensions[header_row].height = 24

    col_widths = [

        len(str(h)) + 4

        for h in headers

    ]

        # =====================================================
    # WRITE DATA
    # =====================================================

    for r_idx, row_values in enumerate(rows):

        excel_row = header_row + 1 + r_idx

        ws.row_dimensions[excel_row].height = 22

        for col_idx, value in enumerate(row_values, start=1):

            cell = ws.cell(
                row=excel_row,
                column=col_idx,
                value=value
            )

            cell.border = CELL_BORDER

            cell.alignment = (
                LEFT
                if col_idx == 1
                else CENTER
            )

            if (
                (col_idx - 1) in number_cols
                and isinstance(value, (int, float))
            ):

                if isinstance(value, float):

                    cell.number_format = "#,##0.00"

                else:

                    cell.number_format = "#,##0"

            if r_idx % 2:

                cell.fill = ZEBRA_FILL

            width = (
                len(str(value)) + 3
                if value is not None
                else 6
            )

            if width > col_widths[col_idx - 1]:

                col_widths[col_idx - 1] = width

        # =====================================================
    # COLUMN WIDTHS
    # =====================================================

    for col_idx, width in enumerate(col_widths, start=1):

        ws.column_dimensions[
            get_column_letter(col_idx)
        ].width = min(max(width, 12), 40)

    # =====================================================
    # TABLE + FILTER
    # =====================================================

    if rows:

        last_row = header_row + len(rows)
        last_col = get_column_letter(len(headers))

        ws.auto_filter.ref = (
            f"A{header_row}:{last_col}{last_row}"
        )

        table = Table(
            displayName=f"Table_{ws.title.replace(' ','_')}",
            ref=f"A{header_row}:{last_col}{last_row}"
        )

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table.tableStyleInfo = style
        ws.add_table(table)

        # ==============================================
        # CONDITIONAL FORMATTING
        # ==============================================

        for c in number_cols:

            col = get_column_letter(c + 1)

            ws.conditional_formatting.add(

                f"{col}{header_row+1}:{col}{last_row}",

                ColorScaleRule(
                    start_type="min",
                    start_color="FFF2CC",

                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFE699",

                    end_type="max",
                    end_color="63BE7B"
                )
            )

    # =====================================================
    # SHEET SETTINGS
    # =====================================================

    ws.freeze_panes = f"B{header_row+1}"

    ws.sheet_view.showGridLines = False

    ws.page_setup.orientation = "landscape"

    ws.page_setup.fitToWidth = 1

    ws.page_margins = PageMargins(
        left=0.30,
        right=0.30,
        top=0.50,
        bottom=0.50
    )

    ws.oddHeader.center.text = "&BMGMS Report"

    ws.oddFooter.left.text = "MGMS"

    ws.oddFooter.center.text = "Confidential"

    ws.oddFooter.right.text = "Page &[Page] of &[Pages]"

    return header_row



def _filters():
    years = [y[0] for y in db.session.query(InformationRow.year).distinct().order_by(InformationRow.year.desc()).all()]
    year = request.args.get("year", type=int) or (years[0] if years else date.today().year)
    month = request.args.get("month", type=int)
    return years, year, month


def _rows(year, month=None):
    q = InformationRow.query.filter_by(year=year)
    if month:
        q = q.filter_by(month=month)
    return q.all()


def _actual_serial(row):
    return (row.gauge_serial or "").strip()


@reports.route("/")
@login_required
def index():
    return redirect(url_for("reports.dashboard"))


def _dashboard_data(year, month):

    rows = _rows(year, month)

    # =====================================================
    # BASIC DATA
    # =====================================================

    actual = [
        (_actual_serial(r), r)
        for r in rows
        if _actual_serial(r)
    ]

    unique_gauges = {
        serial for serial, _ in actual
    }

    wells_set = {
        r.well_number.strip()
        for r in rows
        if r.well_number and r.well_number.strip()
    }

    groups = {
        (r.year, r.month, r.group_no)
        for r in rows
    }

    changes = sum(
        1
        for r in rows
        if r.changed_to and r.changed_to.strip()
    )

    total_days = sum(
        float(r.days or 0)
        for r in rows
    )

    total_jobs = len(groups)

    avg_job_duration = (
        round(total_days / total_jobs, 2)
        if total_jobs
        else 0
    )

    # =====================================================
    # MONTHLY ANALYSIS
    # =====================================================

    jobs_by_month = [0] * 12
    days_by_month = [0.0] * 12
    gauges_by_month = [0] * 12
    wells_by_month = [0] * 12
    changes_by_month = [0] * 12

    monthly_table = []

    for m in range(1, 13):

        mr = [
            r for r in rows
            if r.month == m
        ]

        month_jobs = {
            r.group_no
            for r in mr
        }

        month_gauges = {
            _actual_serial(r)
            for r in mr
            if _actual_serial(r)
        }

        month_wells = {
            r.well_number.strip()
            for r in mr
            if r.well_number
            and r.well_number.strip()
        }

        month_changes = sum(
            1
            for r in mr
            if r.changed_to
            and r.changed_to.strip()
        )

        month_days = sum(
            float(r.days or 0)
            for r in mr
        )

        jobs_count = len(month_jobs)

        jobs_by_month[m - 1] = jobs_count
        days_by_month[m - 1] = round(month_days, 2)
        gauges_by_month[m - 1] = len(month_gauges)
        wells_by_month[m - 1] = len(month_wells)
        changes_by_month[m - 1] = month_changes

        monthly_table.append({
            "month": MONTHS[m - 1],
            "jobs": jobs_count,
            "days": round(month_days, 2),
            "gauges": len(month_gauges),
            "wells": len(month_wells),
            "changes": month_changes,
            "avg_days": (
                round(month_days / jobs_count, 2)
                if jobs_count
                else 0
            )
        })

    # =====================================================
    # GAUGE UTILIZATION
    # =====================================================

    gauge_days = defaultdict(float)

    for serial, r in actual:
        gauge_days[serial] += float(r.days or 0)

    top_gauges = sorted(
        gauge_days.items(),
        key=lambda x: x[1],
        reverse=True
    )[:10]

    gauge_labels = [
        item[0]
        for item in top_gauges
    ]

    gauge_values = [
        round(item[1], 2)
        for item in top_gauges
    ]

    # =====================================================
    # SURVEY ANALYSIS
    # =====================================================

    survey_counts = defaultdict(set)

    for r in rows:

        survey = (
            r.survey.strip()
            if r.survey and r.survey.strip()
            else "Not Specified"
        )

        survey_counts[survey].add(
            (r.year, r.month, r.group_no)
        )

    survey_data = sorted(
        [
            (name, len(job_set))
            for name, job_set in survey_counts.items()
        ],
        key=lambda x: x[1],
        reverse=True
    )

    survey_labels = [
        x[0] for x in survey_data
    ]

    survey_values = [
        x[1] for x in survey_data
    ]

    # =====================================================
    # JOB TYPE ANALYSIS
    # =====================================================

position_counts = defaultdict(set)

for r in rows:

    position = (
        r.position.strip()
        if r.position and r.position.strip()
        else "Not Specified"
    )

    position_counts[position].add(
        (r.year, r.month, r.group_no)
    )

position_data = sorted(
    [
        (name, len(job_set))
        for name, job_set in position_counts.items()
    ],
    key=lambda x: x[1],
    reverse=True
)

    type_labels = [x[0] for x in position_data]
    type_values = [x[1] for x in position_data]

    # =====================================================
    # RIG ANALYSIS
    # =====================================================

    rig_counts = defaultdict(set)

    for r in rows:

        if not r.rig_name or not r.rig_name.strip():
            continue

        rig_counts[r.rig_name.strip()].add(
            (r.year, r.month, r.group_no)
        )

    rig_data = sorted(
        [
            (name, len(job_set))
            for name, job_set in rig_counts.items()
        ],
        key=lambda x: x[1],
        reverse=True
    )[:10]

    rig_labels = [
        x[0] for x in rig_data
    ]

    rig_values = [
        x[1] for x in rig_data
    ]

    # =====================================================
    # WELL ANALYSIS
    # =====================================================

    well_counts = defaultdict(set)

    for r in rows:

        if not r.well_number or not r.well_number.strip():
            continue

        well_counts[r.well_number.strip()].add(
            (r.year, r.month, r.group_no)
        )

    top_wells = sorted(
        [
            (name, len(job_set))
            for name, job_set in well_counts.items()
        ],
        key=lambda x: x[1],
        reverse=True
    )[:10]

    well_labels = [
        x[0] for x in top_wells
    ]

    well_values = [
        x[1] for x in top_wells
    ]

    # =====================================================
    # INSIGHTS
    # =====================================================

    most_active_month = None

    if any(jobs_by_month):

        index = jobs_by_month.index(
            max(jobs_by_month)
        )

        most_active_month = {
            "month": MONTHS[index],
            "jobs": jobs_by_month[index]
        }

    most_used_gauge = (
        top_gauges[0]
        if top_gauges
        else None
    )

    most_active_rig = (
        rig_data[0]
        if rig_data
        else None
    )

    return {
        "total_jobs": total_jobs,
        "total_days": round(total_days, 2),
        "actual_gauges": len(unique_gauges),
        "wells": len(wells_set),
        "changes": changes,
        "avg_job_duration": avg_job_duration,
        "jobs_by_month": jobs_by_month,
        "days_by_month": days_by_month,
        "gauges_by_month": gauges_by_month,
        "wells_by_month": wells_by_month,
        "changes_by_month": changes_by_month,
        "monthly_table": monthly_table,
        "gauge_labels": gauge_labels,
        "gauge_values": gauge_values,
        "survey_labels": survey_labels,
        "survey_values": survey_values,
        "type_labels": type_labels,
        "type_values": type_values,
        "rig_labels": rig_labels,
        "rig_values": rig_values,
        "well_labels": well_labels,
        "well_values": well_values,
        "top_gauges": top_gauges,
        "top_wells": top_wells,
        "most_active_month": most_active_month,
        "most_used_gauge": most_used_gauge,
        "most_active_rig": most_active_rig,
    }


@reports.route("/dashboard")
@login_required
def dashboard():

    years, year, month = _filters()
    d = _dashboard_data(year, month)

    return render_template(
        "reports/dashboard.html",
        years=years,
        year=year,
        month=month,
        months=MONTHS,
        **d
    )

def _actual_gauges_data(year, month):

    rows = _rows(year, month)

    gauges = {}

    for r in rows:

        serial = _actual_serial(r)

        if not serial:
            continue

        if serial not in gauges:

            gauges[serial] = {

                "serial": serial,

                "jobs": set(),

                "wells": set(),

                "changes": 0,

                "days": 0.0,

                "months": [0] * 12,

                "history": []
            }

        g = gauges[serial]

        g["jobs"].add(
            (r.year, r.month, r.group_no)
        )

        if r.well_number and r.well_number.strip():

            g["wells"].add(
                r.well_number.strip()
            )

        days = float(r.days or 0)

        g["days"] += days

        if 1 <= r.month <= 12:

            g["months"][r.month - 1] += days

        if r.changed_to and r.changed_to.strip():

            g["changes"] += 1

        g["history"].append({

            "from_date": r.from_date,

            "to_date": r.to_date,

            "days": round(days, 2),

            "well_number": r.well_number,

            "survey": r.survey,

            "type": r.type,

            "rig_name": r.rig_name,

            "original_gauge": r.gauge_serial,

            "changed_to": r.changed_to,

            "actual_gauge": serial,

            "battery_sn": r.battery_sn,

            "bundle_carrier_sn": r.bundle_carrier_sn,

            "total_hours": round(float(r.total_hours or 0), 2),

            "engineer": r.engineer
        })

    result = []

    for g in gauges.values():

        g["history"].sort(

            key=lambda x: (

                x["from_date"] or date.min,

                x["to_date"] or date.min

            ),

            reverse=True

        )

        result.append({

            "serial": g["serial"],

            "jobs_count": len(g["jobs"]),

            "wells_count": len(g["wells"]),

            "changes": g["changes"],

            "days": round(g["days"], 2),

            "months": [

                round(v, 2)

                for v in g["months"]

            ],

            "history": g["history"]

        })

    result.sort (

        key=lambda x: x["days"],
        reverse=True

    )

    return result

@reports.route("/actual-gauges")
@login_required
def actual_gauges():
    years, year, month = _filters()
    result = _actual_gauges_data(year, month)

    return render_template(
        "reports/actual_gauges.html",
        years=years,
        year=year,
        month=month,
        months=MONTHS,
        gauges=result
    )

def _battery_data(year, month):

    rows = _rows(year, month)

    # =====================================================
    # COLLECT BATTERY USAGE FROM INFORMATION
    # =====================================================

    usage = defaultdict(lambda: {
        "hours": 0.0,
        "days": 0.0,
        "samples": 0,
        "jobs": set(),
        "gauges": set(),
        "history": []
    })

    for r in rows:

        battery_sn = (r.battery_sn or "").strip()

        if not battery_sn:
            continue

        item = usage[battery_sn]

        hours = float(r.total_hours or 0)
        days = float(r.days or 0)
        samples = int(r.total_samples or 0)

        # Total usage from Information
        item["hours"] += hours
        item["days"] += days
        item["samples"] += samples

        # Jobs
        item["jobs"].add(
            (r.year, r.month, r.group_no)
        )

        # Gauge actually used
        gauge_sn = (r.gauge_serial or "").strip()

        if gauge_sn:
            item["gauges"].add(gauge_sn)

        # Usage history
        item["history"].append({

            "year": r.year,
            "month": r.month,
            "group_no": r.group_no,

            "from_date": r.from_date,
            "to_date": r.to_date,

            "well_number": r.well_number,
            "gauge_serial": gauge_sn,

            "hours": hours,
            "days": days,
            "samples": samples,

            "survey": r.survey,
            "type": r.type,
            "rig_name": r.rig_name,
            "engineer": r.engineer
        })

    # =====================================================
    # BUILD BATTERY ANALYSIS
    # =====================================================

    battery_rows = []

    for b in Battery.query.order_by(
        Battery.serial_number
    ).all():

        u = usage[b.serial_number]

        kind = (
            b.compatible_gauge_type or ""
        ).upper()

        capacity = float(b.capacity or 0)

        unit = (
            b.capacity_unit or ""
        ).strip()

        # Consumption before MGMS
        previous_consumption = float(
            b.previous_consumption or 0
        )

        # Consumption calculated from Information
        mgms_consumption = None

        # Previous + MGMS
        total_consumed = None

        formula = ""

        # =================================================
        # EGYWELL
        # =================================================

        if "EGY" in kind:

            mgms_consumption = (
                (u["hours"] * 0.05)
                +
                (u["samples"] * 5 / 3600)
            )

            formula = (
                "Hours × 0.05 + Samples × 5 / 3600"
            )

        # =================================================
        # JABERTEK
        # =================================================

        elif "JAB" in kind:

            h = u["hours"]

            if h <= 0:

                mgms_consumption = 0

            elif h <= 1:

                mgms_consumption = 0.0064

            else:

                mgms_consumption = (
                    0.0064
                    +
                    ((h - 1) * 0.00445)
                )

            formula = (
                "First hour @10 sec = 0.0064 Ah; "
                "remaining hours × 0.00445"
            )

        # =================================================
        # METROLOG
        # =================================================

        elif "MET" in kind:

            mgms_consumption = u["days"]

            formula = "Actual working days"

        # =================================================
        # UNKNOWN
        # =================================================

        else:

            mgms_consumption = None

            formula = (
                "Calculation method not configured"
            )

        # =================================================
        # TOTAL CONSUMED
        # =================================================

        if mgms_consumption is not None:

            total_consumed = (
                previous_consumption
                +
                mgms_consumption
            )

        # =================================================
        # REMAINING / PERCENTAGES
        # =================================================

        remaining = None
        used_percent = None
        remaining_percent = None

        if (
            total_consumed is not None
            and capacity > 0
        ):

            remaining = max(
                capacity - total_consumed,
                0
            )

            used_percent = min(
                (total_consumed / capacity) * 100,
                100
            )

            remaining_percent = max(
                100 - used_percent,
                0
            )

        # =================================================
        # SORT HISTORY
        # =================================================

        u["history"].sort(
            key=lambda x: (
                x["from_date"] or date.min,
                x["group_no"] or 0
            ),
            reverse=True
        )

        # =================================================
        # FINAL DATA
        # =================================================

        battery_rows.append({

            "battery": b,

            # -------------------------
            # Actual usage
            # -------------------------

            "hours": round(
                u["hours"],
                2
            ),

            "days": round(
                u["days"],
                2
            ),

            "samples": u["samples"],

            "jobs_count": len(
                u["jobs"]
            ),

            "gauges_count": len(
                u["gauges"]
            ),

            "gauges": sorted(
                u["gauges"]
            ),

            # -------------------------
            # Capacity
            # -------------------------

            "capacity": round(
                capacity,
                4
            ),

            "unit": unit,

            # -------------------------
            # Previous consumption
            # -------------------------

            "previous_consumption": round(
                previous_consumption,
                4
            ),

            # -------------------------
            # MGMS consumption
            # -------------------------

            "mgms_consumption": (
                None
                if mgms_consumption is None
                else round(mgms_consumption, 4)
            ),

            # -------------------------
            # Total consumed
            # -------------------------

            "total_consumed": (
                None
                if total_consumed is None
                else round(total_consumed, 4)
            ),

            # Keep old key temporarily
            # so current HTML doesn't break
            "consumed": (
                None
                if total_consumed is None
                else round(total_consumed, 4)
            ),

            # -------------------------
            # Remaining
            # -------------------------

            "remaining": (
                None
                if remaining is None
                else round(remaining, 4)
            ),

            # -------------------------
            # Percentages
            # -------------------------

            "used_percent": (
                None
                if used_percent is None
                else round(used_percent, 2)
            ),

            "remaining_percent": (
                None
                if remaining_percent is None
                else round(remaining_percent, 2)
            ),

            # -------------------------
            # Formula / History
            # -------------------------

            "formula": formula,

            "history": u["history"]
        })

    return battery_rows


@reports.route("/batteries")
@login_required
def batteries():

    years, year, month = _filters()
    battery_rows = _battery_data(year, month)

    return render_template(
        "reports/batteries.html",
        years=years,
        year=year,
        month=month,
        months=MONTHS,
        batteries=battery_rows
    )






@reports.route("/maintenance", methods=["GET", "POST"])
@login_required
def maintenance():

    if request.method == "POST":

        maintenance_id = request.form.get("maintenance_id")

        if maintenance_id:

            record = MaintenanceRecord.query.get_or_404(maintenance_id)

        else:

            record = MaintenanceRecord()

        record.equipment_type = request.form.get("equipment_type", "").strip()
        record.serial_number = request.form.get("serial_number", "").strip()
        record.maintenance_date = datetime.strptime(
            request.form["maintenance_date"],
            "%Y-%m-%d"
        ).date()

        record.problem = request.form.get("problem", "").strip()
        record.action_taken = request.form.get("action_taken", "").strip()
        record.status = request.form.get("status", "Maintenance")

        record.return_date = (
            datetime.strptime(
                request.form["return_date"],
                "%Y-%m-%d"
            ).date()
            if request.form.get("return_date")
            else None
        )

        record.notes = request.form.get("notes", "").strip()

        if not maintenance_id:
            db.session.add(record)

        db.session.commit()

        flash(
            "Maintenance record updated."
            if maintenance_id
            else "Maintenance record added.",
            "success"
        )

        return redirect(url_for("reports.maintenance"))

    records = MaintenanceRecord.query.order_by(
        MaintenanceRecord.maintenance_date.desc(),
        MaintenanceRecord.id.desc()
    ).all()

    return render_template(
        "reports/maintenance.html",
        records=records
    )
@reports.route("/maintenance/<int:maintenance_id>/delete", methods=["POST"])
@login_required
def delete_maintenance(maintenance_id):

    record = MaintenanceRecord.query.get_or_404(maintenance_id)

    db.session.delete(record)
    db.session.commit()

    flash("Maintenance record deleted.", "success")

    return redirect(url_for("reports.maintenance"))

@reports.route("/performance")
@login_required
def performance():
    years, year, month = _filters()
    rows = _rows(year, month)
    stats = defaultdict(lambda: {"days": 0.0, "jobs": set(), "changes": 0})
    for r in rows:
        sn = _actual_serial(r)
        if sn:
            stats[sn]["days"] += float(r.days or 0)
            stats[sn]["jobs"].add((r.year, r.month, r.group_no))
            stats[sn]["changes"] += 1 if r.changed_to and r.changed_to.strip() else 0
    maint = defaultdict(int)
    for m in MaintenanceRecord.query.filter_by(equipment_type="Gauge").all():
        maint[m.serial_number] += 1
    items = [{"serial": k, "days": round(v["days"],2), "jobs": len(v["jobs"]), "changes": v["changes"], "maintenance": maint[k]} for k,v in stats.items()]
    items.sort(key=lambda x: x["days"], reverse=True)
    return render_template(
    "reports/performance.html",
    years=years,
    year=year,
    month=month,
    months=MONTHS,
    items=items
)


@reports.route("/job-history")
@login_required
def job_history():
    q = request.args.get("q", "").strip()
    query = InformationRow.query
    if q:
        like = f"%{q}%"
        query = query.filter(db.or_(InformationRow.well_number.ilike(like), InformationRow.gauge_serial.ilike(like), InformationRow.changed_to.ilike(like), InformationRow.battery_sn.ilike(like), InformationRow.bundle_carrier_sn.ilike(like)))
    rows = query.order_by(InformationRow.year.desc(), InformationRow.month.desc(), InformationRow.group_no.desc()).limit(300).all()
    return render_template("reports/job_history.html", rows=rows, q=q)


@reports.route("/engineers")
@login_required
def engineers():

    years, year, month = _filters()
    rows = _rows(year, month)

    stats = defaultdict(lambda: {
        "jobs": set(),
        "days": 0.0,
        "hours": 0.0,
        "wells": set(),
        "gauges": set(),
        "changes": 0,
        "months": [0.0] * 12
    })

    for r in rows:

        name = (r.engineer or "").strip()

        if not name:
            continue

        s = stats[name]

        s["jobs"].add((r.year, r.month, r.group_no))
        s["days"] += float(r.days or 0)
        s["hours"] += float(r.total_hours or 0)

        if r.well_number and r.well_number.strip():
            s["wells"].add(r.well_number.strip())

        gauge_sn = _actual_serial(r)
        if gauge_sn:
            s["gauges"].add(gauge_sn)

        if r.changed_to and r.changed_to.strip():
            s["changes"] += 1

        if r.month and 1 <= r.month <= 12:
            s["months"][r.month - 1] += float(r.total_hours or 0)

    engineer_rows = []

    for name, s in stats.items():

        engineer_rows.append({
            "name": name,
            "jobs": len(s["jobs"]),
            "days": round(s["days"], 2),
            "hours": round(s["hours"], 2),
            "wells": len(s["wells"]),
            "gauges": len(s["gauges"]),
            "changes": s["changes"],
            "avg_hours_per_job": (
                round(s["hours"] / len(s["jobs"]), 2)
                if s["jobs"] else 0
            ),
            "months": [round(v, 2) for v in s["months"]]
        })

    engineer_rows.sort(key=lambda x: x["hours"], reverse=True)

    total_engineers = len(engineer_rows)
    total_jobs = sum(e["jobs"] for e in engineer_rows)
    total_hours = round(sum(e["hours"] for e in engineer_rows), 2)

    avg_hours_per_engineer = (
        round(total_hours / total_engineers, 2)
        if total_engineers else 0
    )

    top_engineers = engineer_rows[:10]

    top_labels = [e["name"] for e in top_engineers]
    top_values = [e["hours"] for e in top_engineers]

    busiest_engineer = engineer_rows[0] if engineer_rows else None

    return render_template(
        "reports/engineers.html",
        years=years,
        year=year,
        month=month,
        months=MONTHS,
        engineers=engineer_rows,
        total_engineers=total_engineers,
        total_jobs=total_jobs,
        total_hours=total_hours,
        avg_hours_per_engineer=avg_hours_per_engineer,
        top_labels=top_labels,
        top_values=top_values,
        busiest_engineer=busiest_engineer
    )


# ==========================================================
#                    EXPORT TO EXCEL
# ==========================================================

@reports.route("/export/<kind>")
@login_required
def export_report(kind):

    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    q = request.args.get("q", "").strip()

    wb = Workbook()
    ws = wb.active

    filename = f"Report_{kind}.xlsx"

    if kind == "dashboard":

        years, resolved_year, resolved_month = _filters()
        target_year = year or resolved_year

        d = _dashboard_data(target_year, month)
        d["year"] = target_year
        report = ExcelReport()
        report.dashboard(d)
        return send_file(
            report.save(),
            as_attachment=True,
             download_name=f"Dashboard_Full_Report_{target_year}.xlsx",
             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        

        # ==========================================
        # New XlsxWriter Dashboard
        # ==========================================
        # ---------------- Sheet 1: Overview (KPIs) ----------------

        ws.title = "Overview"

        kpi_rows = [
            ("Total Jobs", d["total_jobs"]),
            ("Total Working Days", d["total_days"]),
            ("Actual Gauges Used", d["actual_gauges"]),
            ("Unique Wells", d["wells"]),
            ("Gauge Changes", d["changes"]),
            ("Avg Job Duration (days)", d["avg_job_duration"]),
            ("Most Active Month", d["most_active_month"]["month"] if d["most_active_month"] else "-"),
            ("Most Used Gauge", d["most_used_gauge"][0] if d["most_used_gauge"] else "-"),
            ("Most Active Rig", d["most_active_rig"][0] if d["most_active_rig"] else "-"),
        ]

        _write_sheet(
            ws,
            ["Metric", "Value"],
            kpi_rows,
            title="Operations Dashboard — Overview",
            subtitle=f"Year {target_year}" + (f" / {MONTHS[month-1]}" if month else " (All Months)")
        )

        # ---------------- Sheet 2: Monthly Analysis + chart ----------------

        ws_monthly = wb.create_sheet("Monthly Analysis")

        monthly_rows = [
            (
                mt["month"], mt["jobs"], mt["days"], mt["gauges"],
                mt["wells"], mt["changes"], mt["avg_days"]
            )
            for mt in d["monthly_table"]
        ]

        header_row = _write_sheet(
            ws_monthly,
            ["Month", "Jobs", "Total Days", "Actual Gauges", "Wells", "Changes", "Avg Days/Job"],
            monthly_rows,
            title="Monthly Analysis",
            number_cols={1, 2, 3, 4, 5, 6}
        )

        chart1 = BarChart()
        chart1.title = "Jobs & Working Days per Month"
        chart1.y_axis.title = "Count"
        chart1.x_axis.title = "Month"
        chart1.height = 9
        chart1.width = 22

        data_ref = Reference(
            ws_monthly, min_col=2, max_col=3,
            min_row=header_row, max_row=header_row + len(monthly_rows)
        )
        cats_ref = Reference(
            ws_monthly, min_col=1,
            min_row=header_row + 1, max_row=header_row + len(monthly_rows)
        )
        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(cats_ref)

        ws_monthly.add_chart(chart1, f"I{header_row}")

        # ---------------- Sheet 3: Gauge Utilization + chart ----------------

        ws_gauges = wb.create_sheet("Gauge Utilization")

        gauge_rows = list(zip(d["gauge_labels"], d["gauge_values"]))

        header_row = _write_sheet(
            ws_gauges,
            ["Gauge", "Working Days"],
            gauge_rows,
            title="Top Gauges by Utilization",
            number_cols={1}
        )

        if gauge_rows:
            chart2 = BarChart()
            chart2.title = "Top Gauges by Working Days"
            chart2.height = 9
            chart2.width = 18

            data_ref = Reference(
                ws_gauges, min_col=2,
                min_row=header_row, max_row=header_row + len(gauge_rows)
            )
            cats_ref = Reference(
                ws_gauges, min_col=1,
                min_row=header_row + 1, max_row=header_row + len(gauge_rows)
            )
            chart2.add_data(data_ref, titles_from_data=True)
            chart2.set_categories(cats_ref)
            ws_gauges.add_chart(chart2, f"D{header_row}")

        # ---------------- Sheet 4: Survey Analysis + pie chart ----------------

        ws_survey = wb.create_sheet("Survey Analysis")

        survey_rows = list(zip(d["survey_labels"], d["survey_values"]))

        header_row = _write_sheet(
            ws_survey,
            ["Survey Type", "Jobs"],
            survey_rows,
            title="Jobs by Survey Type",
            number_cols={1}
        )

        if survey_rows:
            pie1 = PieChart()
            pie1.title = "Survey Type Distribution"
            pie1.height = 9
            pie1.width = 14

            data_ref = Reference(
                ws_survey, min_col=2,
                min_row=header_row, max_row=header_row + len(survey_rows)
            )
            cats_ref = Reference(
                ws_survey, min_col=1,
                min_row=header_row + 1, max_row=header_row + len(survey_rows)
            )
            pie1.add_data(data_ref, titles_from_data=True)
            pie1.set_categories(cats_ref)
            ws_survey.add_chart(pie1, f"D{header_row}")

        # ---------------- Sheet 5: Job Type Analysis + pie chart ----------------

        ws_type = wb.create_sheet("Job Type Analysis")

        type_rows = list(zip(d["type_labels"], d["type_values"]))

        header_row = _write_sheet(
            ws_type,
            ["Job Type", "Jobs"],
            type_rows,
            title="Jobs by Type",
            number_cols={1}
        )

        if type_rows:
            pie2 = PieChart()
            pie2.title = "Job Type Distribution"
            pie2.height = 9
            pie2.width = 14

            data_ref = Reference(
                ws_type, min_col=2,
                min_row=header_row, max_row=header_row + len(type_rows)
            )
            cats_ref = Reference(
                ws_type, min_col=1,
                min_row=header_row + 1, max_row=header_row + len(type_rows)
            )
            pie2.add_data(data_ref, titles_from_data=True)
            pie2.set_categories(cats_ref)
            ws_type.add_chart(pie2, f"D{header_row}")

        # ---------------- Sheet 6: Rig Analysis + chart ----------------

        ws_rig = wb.create_sheet("Rig Analysis")

        rig_rows = list(zip(d["rig_labels"], d["rig_values"]))

        header_row = _write_sheet(
            ws_rig,
            ["Rig", "Jobs"],
            rig_rows,
            title="Top Rigs by Job Count",
            number_cols={1}
        )

        if rig_rows:
            chart3 = BarChart()
            chart3.title = "Top Rigs by Jobs"
            chart3.height = 9
            chart3.width = 18

            data_ref = Reference(
                ws_rig, min_col=2,
                min_row=header_row, max_row=header_row + len(rig_rows)
            )
            cats_ref = Reference(
                ws_rig, min_col=1,
                min_row=header_row + 1, max_row=header_row + len(rig_rows)
            )
            chart3.add_data(data_ref, titles_from_data=True)
            chart3.set_categories(cats_ref)
            ws_rig.add_chart(chart3, f"D{header_row}")

        # ---------------- Sheet 7: Well Analysis + chart ----------------

        ws_well = wb.create_sheet("Well Analysis")

        well_rows = list(zip(d["well_labels"], d["well_values"]))

        header_row = _write_sheet(
            ws_well,
            ["Well", "Jobs"],
            well_rows,
            title="Top Wells by Job Count",
            number_cols={1}
        )

        if well_rows:
            chart4 = BarChart()
            chart4.title = "Top Wells by Jobs"
            chart4.height = 9
            chart4.width = 18

            data_ref = Reference(
                ws_well, min_col=2,
                min_row=header_row, max_row=header_row + len(well_rows)
            )
            cats_ref = Reference(
                ws_well, min_col=1,
                min_row=header_row + 1, max_row=header_row + len(well_rows)
            )
            chart4.add_data(data_ref, titles_from_data=True)
            chart4.set_categories(cats_ref)
            ws_well.add_chart(chart4, f"D{header_row}")

        filename = f"Dashboard_Full_Report_{target_year}.xlsx"

    elif kind == "actual-gauges":

        years, resolved_year, resolved_month = _filters()

        target_year = year or resolved_year
        target_month = month if month is not None else resolved_month

        gauges = _actual_gauges_data(
            target_year,
            target_month
        )

        ws.title = "Summary"

        summary_rows = []

        for g in gauges:

            summary_rows.append(

                (
                    g["serial"],
                    g["jobs_count"],
                    g["wells_count"],
                    g["changes"],

                    g["months"][0],
                    g["months"][1],
                    g["months"][2],
                    g["months"][3],
                    g["months"][4],
                    g["months"][5],
                    g["months"][6],
                    g["months"][7],
                    g["months"][8],
                    g["months"][9],
                    g["months"][10],
                    g["months"][11],

                    g["days"]
                )

            )

        _write_sheet(

            ws,

            [
                "Gauge",
                "Jobs",
                "Wells",
                "Changes",

                "Jan",
                "Feb",
                "Mar",
                "Apr",
                "May",
                "Jun",
                "Jul",
                "Aug",
                "Sep",
                "Oct",
                "Nov",
                "Dec",

                "Total Days"
            ],

            summary_rows,

            title="Actual Gauges — Summary",

            subtitle=(
                f"Year {target_year}"
                + (
                    f" / {MONTHS[target_month-1]}"
                    if target_month
                    else " (All Months)"
                )
            ),

            number_cols={
                1, 2, 3,
                4, 5, 6, 7,
                8, 9, 10, 11,
                12, 13, 14, 15,
                16
            }

        )

        ws2 = wb.create_sheet("History")

        history_rows = []

        for g in gauges:

            for h in g["history"]:

                history_rows.append(

                    (

                        h["from_date"].strftime("%d %b %Y") if h["from_date"] else "",

                        h["to_date"].strftime("%d %b %Y") if h["to_date"] else "",

                        h["days"],

                        h["well_number"] or "",

                        h["survey"] or "",

                        h["type"] or "",

                        h["rig_name"] or "",

                        h["original_gauge"] or "",

                        h["changed_to"] or "",

                        h["actual_gauge"] or "",

                        h["battery_sn"] or "",

                        h["bundle_carrier_sn"] or "",

                        h["total_hours"],

                        h["engineer"] or ""

                    )

                )

        _write_sheet(

            ws2,

            [

                "From",
                "To",
                "Days",
                "Well",
                "Survey",
                "Type",
                "Rig",
                "Original Gauge",
                "Changed To",
                "Actual Gauge",
                "Battery",
                "Bundle",
                "Hours",
                "Engineer"

            ],

            history_rows,

            title="Actual Gauges — Full Work History",

            number_cols={2, 12}

        )

        filename = f"Actual_Gauges_{target_year}.xlsx"
    elif kind == "batteries":
        years, resolved_year, resolved_month = _filters()

        target_year = year or resolved_year
        target_month = month if month is not None else resolved_month

        batteries = _battery_data(target_year, target_month)

        ws.title = "Batteries"
        table = []
        for b in batteries:
            table.append(
                (
                    b["battery"].serial_number,
                    b["battery"].compatible_gauge_type,
                    b["capacity"],
                    b["unit"],
                    b["previous_consumption"],
                    b["mgms_consumption"],
                    b["total_consumed"],
                    b["remaining"],
                    b["used_percent"],
                    b["remaining_percent"],
                    b["jobs_count"],
                    b["gauges_count"]
                )
            )

        _write_sheet(
            ws,
            [
                "Battery",
                "Gauge Type",
                "Capacity",
                "Unit",
                "Previous",
                "MGMS",
                "Consumed",
                "Remaining",
                "Used %",
                "Remaining %",
                "Jobs",
                "Gauges"
            ],
            table,
            title="Battery Consumption Report",
            subtitle=(
                f"Year {target_year}"
                + (
                    f" / {MONTHS[target_month-1]}"
                    if target_month
                    else " (All Months)"
                )
            ),
            number_cols={2, 3, 4, 5, 6, 7, 8, 9, 10, 11}
        )

        filename = f"Battery_Report_{target_year}.xlsx"
    elif kind == "performance":

        years, resolved_year, resolved_month = _filters()
        target_year = year or resolved_year

        rows = _rows(target_year, month)

        stats = defaultdict(lambda: {"days": 0.0, "jobs": set(), "changes": 0})

        for r in rows:
            sn = _actual_serial(r)
            if sn:
                stats[sn]["days"] += float(r.days or 0)
                stats[sn]["jobs"].add((r.year, r.month, r.group_no))
                stats[sn]["changes"] += 1 if r.changed_to and r.changed_to.strip() else 0

        maint = defaultdict(int)
        for m in MaintenanceRecord.query.filter_by(equipment_type="Gauge").all():
            maint[m.serial_number] += 1

        table = [
            (k, len(v["jobs"]), round(v["days"], 2), v["changes"], maint[k])
            for k, v in stats.items()
        ]
        table.sort(key=lambda x: x[2], reverse=True)

        ws.title = "Performance"

        _write_sheet(
            ws,
            ["Gauge", "Jobs", "Working Days", "Changes", "Maintenance Records"],
            table,
            title="Equipment Performance Report",
            subtitle=f"Year {target_year}" + (f" / {MONTHS[month-1]}" if month else " (All Months)"),
            number_cols={1, 2, 3, 4}
        )

        filename = f"Equipment_Performance_{target_year}.xlsx"

    elif kind == "job-history":

        query = InformationRow.query

        if q:
            like = f"%{q}%"
            query = query.filter(db.or_(
                InformationRow.well_number.ilike(like),
                InformationRow.gauge_serial.ilike(like),
                InformationRow.changed_to.ilike(like),
                InformationRow.battery_sn.ilike(like),
                InformationRow.bundle_carrier_sn.ilike(like)
            ))

        rows = query.order_by(
            InformationRow.year.desc(),
            InformationRow.month.desc(),
            InformationRow.group_no.desc()
        ).limit(300).all()

        ws.title = "Job History"

        table = [
            (
                f"{r.month}/{r.year}",
                r.well_number,
                r.gauge_serial,
                r.changed_to or "",
                r.days,
                r.total_hours,
                r.battery_sn or "",
                r.bundle_carrier_sn or "",
                r.survey or ""
            )
            for r in rows
        ]

        _write_sheet(
            ws,
            ["Period", "Well", "Gauge Used", "Changed To", "Days",
             "Hours", "Battery", "Bundle", "Survey"],
            table,
            title="Job History",
            subtitle=f"Search: \"{q}\"" if q else "All records (latest 300)",
            number_cols={4, 5}
        )

        filename = "Job_History.xlsx"

    elif kind == "maintenance":

        records = MaintenanceRecord.query.order_by(
            MaintenanceRecord.maintenance_date.desc(),
            MaintenanceRecord.id.desc()
        ).all()

        ws.title = "Maintenance"

        table = [
            (
                r.maintenance_date.strftime("%d %b %Y") if r.maintenance_date else "",
                r.equipment_type,
                r.serial_number,
                r.problem,
                r.action_taken or "",
                r.status,
                r.return_date.strftime("%d %b %Y") if r.return_date else ""
            )
            for r in records
        ]

        _write_sheet(
            ws,
            ["Date", "Equipment", "Serial", "Problem", "Action", "Status", "Return"],
            table,
            title="Maintenance Report",
            subtitle="All equipment maintenance records"
        )

        filename = "Maintenance_Report.xlsx"

    elif kind == "engineers":

        years, resolved_year, resolved_month = _filters()
        target_year = year or resolved_year

        rows = _rows(target_year, month)

        stats = defaultdict(lambda: {
            "jobs": set(), "days": 0.0, "hours": 0.0,
            "wells": set(), "gauges": set(), "changes": 0
        })

        for r in rows:

            name = (r.engineer or "").strip()
            if not name:
                continue

            s = stats[name]
            s["jobs"].add((r.year, r.month, r.group_no))
            s["days"] += float(r.days or 0)
            s["hours"] += float(r.total_hours or 0)

            if r.well_number and r.well_number.strip():
                s["wells"].add(r.well_number.strip())

            gauge_sn = _actual_serial(r)
            if gauge_sn:
                s["gauges"].add(gauge_sn)

            if r.changed_to and r.changed_to.strip():
                s["changes"] += 1

        table = [
            (
                name, len(s["jobs"]), round(s["days"], 2), round(s["hours"], 2),
                len(s["wells"]), len(s["gauges"]), s["changes"]
            )
            for name, s in stats.items()
        ]
        table.sort(key=lambda x: x[3], reverse=True)

        ws.title = "Engineers"

        header_row = _write_sheet(
            ws,
            ["Engineer", "Jobs", "Working Days", "Total Hours", "Wells", "Gauges Used", "Changes"],
            table,
            title="Engineers Analytics Report",
            subtitle=f"Year {target_year}" + (f" / {MONTHS[month-1]}" if month else " (All Months)"),
            number_cols={1, 2, 3, 4, 5, 6}
        )

        if table:
            chart5 = BarChart()
            chart5.title = "Total Hours by Engineer"
            chart5.height = 9
            chart5.width = 20

            data_ref = Reference(
                ws, min_col=4,
                min_row=header_row, max_row=header_row + len(table)
            )
            cats_ref = Reference(
                ws, min_col=1,
                min_row=header_row + 1, max_row=header_row + len(table)
            )
            chart5.add_data(data_ref, titles_from_data=True)
            chart5.set_categories(cats_ref)
            ws.add_chart(chart5, f"I{header_row}")

        filename = f"Engineers_Report_{target_year}.xlsx"
    else:

        return "Unknown report type", 400

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
