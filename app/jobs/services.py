import calendar
from datetime import date

from sqlalchemy import func

from app.extensions import db
from app.models.jobs import Job
from app.models.memory_gauge import MemoryGauge
from app.models.bundle_carrier import BundleCarrier


def get_worked_days(equipment_type, equipment_id, month, year):
    """
    ترجع عدد الأيام الفريدة التي عمل فيها الجهاز خلال الشهر.
    """

    return (
        db.session.query(func.count(func.distinct(Job.day)))
        .filter(
            Job.equipment_type == equipment_type,
            Job.equipment_id == equipment_id,
            Job.month == month,
            Job.year == year,
        )
        .scalar()
        or 0
    )


def get_month_days(year, month):
    """
    ترجع لستة بعدد أيام الشهر الفعلي (28-31)، كل عنصر فيها
    dict فيه رقم اليوم واسم اليوم (Mon, Tue...).
    """

    days_in_month = calendar.monthrange(year, month)[1]

    result = []

    for day in range(1, days_in_month + 1):
        weekday_name = date(year, month, day).strftime("%a")

        result.append({
            "day": day,
            "weekday": weekday_name,
        })

    return result


def build_jobs_workbook(year, month):
    """
    تبني ملف Excel فيه شبكة الجوبز بنفس شكل الجدول:
    الصفوف = أيام الشهر، الأعمدة = الجوجات والباندل كاريرز.
    """

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    gauges = MemoryGauge.query.order_by(MemoryGauge.serial_number).all()
    bundles = BundleCarrier.query.order_by(BundleCarrier.serial_number).all()

    jobs = Job.query.filter_by(year=year, month=month).all()

    jobs_map = {}
    for job in jobs:
        jobs_map[(job.day, job.equipment_type, job.equipment_id)] = job.well_number

    month_days = get_month_days(year, month)
    month_name = calendar.month_name[month]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {year}"[:31]

    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill("solid", fgColor="38B000")
    thin_border = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
    center = Alignment(horizontal="center", vertical="center")

    # ---- Header row ----
    headers = ["Date", "Day"] + [g.serial_number for g in gauges] + [b.serial_number for b in bundles]

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # ---- Data rows ----
    for row_idx, item in enumerate(month_days, start=2):
        day = item["day"]

        ws.cell(row=row_idx, column=1, value=f"{day:02d} {month_name} {year}").border = thin_border
        ws.cell(row=row_idx, column=2, value=item["weekday"]).border = thin_border

        col = 3
        for gauge in gauges:
            value = jobs_map.get((day, "gauge", gauge.id), "")
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = center
            cell.border = thin_border
            col += 1

        for bundle in bundles:
            value = jobs_map.get((day, "bundle", bundle.id), "")
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = center
            cell.border = thin_border
            col += 1

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 8

    for i in range(3, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 14

    ws.freeze_panes = "C2"

    return wb
